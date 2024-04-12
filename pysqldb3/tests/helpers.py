import random
import os
import pandas as pd
import requests
import zipfile
import configparser
import csv
import openpyxl
from xlrd import open_workbook
from xlutils.copy import copy
from zipfile import ZipFile
from ..Config import write_config
write_config(confi_path=os.path.dirname(os.path.abspath(__file__)).replace('\\tests','') + "\\config.cfg")

DIR = os.path.join(os.path.dirname(os.path.abspath(__file__))) + '\\test_data'


def set_up_test_csv():
    data = {'id': {0: 1, 1: 2, 2: 3, 3: 4, 4: 5},
            'living space (sq ft)': {0: 2222, 1: 1628, 2: 3824, 3: 1137, 4: 3560},
            'beds': {0: 3, 1: 3, 2: 5, 3: 3, 4: 6},
            'baths': {0: 3, 1: 2.5, 2: 4.0, 3: 2.0, 4: 4.0},
            'zip': {0: 32312, 1: 32308, 2: 32312, 3: 32309, 4: 32309},
            'year': {0: 1981, 1: 2009, 2: 1954, 3: 1993, 4: 1973},
            'list price': {0: 250000, 1: 185000, 2: 399000, 3: 150000, 4: 315000},
            'neighborhood': {0: 'Corona', 1: 'Kensington', 2: 'Morris Heights', 3: 'Bayside', 4: 'Inwood'},
            'sale date': {0: '1/1/2015', 1: '2/25/2012', 2: '7/9/2018', 3: '12/2/2021', 4: '11/13/1995'}}
    df = pd.DataFrame(data)
    df.to_csv(DIR+"\\test.csv", index=False)
    df.to_csv(DIR+"\\test2.csv", index=False, sep='|')
    df.to_csv(DIR+"\\test3.csv", index=False, header=False)

    data['neighborhood'][0] = data['neighborhood'][0] * 500
    df.to_csv(DIR+"\\varchar.csv", index=False, header=False)

    # add csv with extra header line
    data2 = [
        ['header to skip', 'header to skip', 'header to skip'],
        ['real_header1', 'real header2', 'real header 3'],
        [1, 2, 3],
        [9, 9, 9]
    ]

    # simple csv with empty column
    df = pd.DataFrame({"id": {1: 1, 2: 2, 3: 3},
                       "col1": {1: 'a', 2: 'b'},
                       "col3": {1: None, 2: None},
                       "col2": {1: 35, 2: 0},
                       })
    df.to_csv(DIR + "\\test6.csv", index=False)

    # bulk csv with empty column
    data3 = [
        ["id",  "col1", "col3", 'col2'],
        [1,2,None,100],
        [2, 3,None, 9],
        [35,36,None, 37]
    ]
    for i in range(1000):
        data3.append([2+i, 2, None, 0])

    with open(DIR+"\\test7.csv", 'w', newline='') as csvfile:
        w = csv.writer(csvfile, delimiter=',')
        for row in data3:
            w.writerow(row)

    with open(DIR+"\\test4.csv", 'w', newline='') as csvfile:
        w = csv.writer(csvfile, delimiter=',')
        for row in data2:
            w.writerow(row)

    # create bulk csv with extra header to skip
    for i in range(1000):
        data2.append([1, 2, 3])

    with open(DIR+"\\test5.csv", 'w', newline='') as csvfile:
        w = csv.writer(csvfile, delimiter=',')
        for row in data2:
            w.writerow(row)


def set_up_test_table_sql(sql, schema='dbo'):
    """
    Creates one test table for testing

    Uses random to make randomly generated inputs.
    """
    table_name = 'sql_test_table_{}'.format(sql.user)

    if sql.table_exists(table=table_name, schema=schema):
        return

    sql.query("""
    create table {s}.{t} (test_col1 int, test_col2 int, geom geometry);
    insert into {s}.{t} VALUES(1, 2, geometry::Point(985831.79200444, 203371.60461367, 2263));
    insert into {s}.{t} VALUES(3, 4, geometry::Point(985831.79200444, 203371.60461367, 2263));
    """.format(s=schema, t=table_name))


def clean_up_test_table_sql(sql, schema='dbo'):
    table_name = 'sql_test_table_{}'.format(sql.user)
    sql.drop_table(table=table_name, schema=schema)


def set_up_test_table_pg(db, schema='working'):
    """
    Creates one test table for testing

    Uses random to make randomly generated inputs.
    """
    table_name = 'pg_test_table_{}'.format(db.user)

    if db.table_exists(table=table_name, schema=schema):
        return

    db.query("""
    create table {}.{}(id int, test_col1 int, test_col2 int, geom geometry);
    """.format(schema, table_name))

    for i in range(0, 1000):
        c1 = random.randint(0, 10000)
        c2 = random.randint(0, 10000)

        dec_lat = random.random() / 100
        dec_lon = random.random() / 100

        lat = 40.7 + dec_lat
        lon = -74 + dec_lon

        db.query("""
        INSERT INTO {}.{} values
        ({}, {}, {}, ST_SetSRID(ST_MakePoint({}, {}), 4326))
        """.format(schema, table_name, i, c1, c2, lat, lon))


def clean_up_test_table_pg(db):
    table_name = 'pg_test_table_{}'.format(db.user)
    db.drop_table(table=table_name, schema='working')


def set_up_two_test_tables_pg(db):
    """
    Creates two test tables for testing

    Uses random to make randomly generated inputs.
    """
    table_name = 'pg_test_table_{}'.format(db.user)
    table_name2 = 'pg_test_table_{}_2'.format(db.user)

    if db.table_exists(table=table_name, schema='working') and \
            db.table_exists(table=table_name2, schema='working'):
        return
    else:
        db.drop_table(table=table_name, schema='working')
        db.drop_table(table=table_name2, schema='working')

    db.query("""
    create table working.{}(id int, test_col1 int, test_col2 int, geom geometry);
    """.format(table_name))

    db.query("""
    create table working.{}(id int, test_col1 int, test_col2 int, geom geometry);
    """.format(table_name2))

    for i in range(0, 10000):
        c1 = random.randint(0, 10000)
        c2 = random.randint(0, 10000)

        dec_lat = random.random() / 100
        dec_lon = random.random() / 100

        lat = 40.7 + dec_lat
        lon = -74 + dec_lon

        dec_lat2 = random.random() / 100
        dec_lon2 = random.random() / 100

        lat2 = 40.7 + dec_lat2
        lon2 = -74 + dec_lon2

        db.query("""
        INSERT INTO working.{} values
        ({}, {}, {}, ST_SetSRID(ST_MakePoint({}, {}), 4326))
        """.format(table_name, i, c1, c2, lat, lon))

        db.query("""
        INSERT INTO working.{} values
        ({}, {}, {}, ST_SetSRID(ST_MakePoint({}, {}), 4326))
        """.format(table_name2, i, c1, c2, lat2, lon2))


def clean_up_two_test_tables_pg(db):
    table_name = 'pg_test_table_{}'.format(db.user)
    table_name2 = 'pg_test_table_{}_2'.format(db.user)

    db.drop_table(table=table_name, schema='working')
    db.drop_table(table=table_name2, schema='working')


def set_up_feature_class():
    """
    Builds file gdb with a feature class with sample data
    :return:
    """
    zip_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'test_data/nyclion_21d.zip')
    if not os.path.isfile(zip_path):

        download_url = r'https://www1.nyc.gov/assets/planning/download/zip/data-maps/open-data/nyclion_21d.zip'

        config = configparser.ConfigParser()
        config.read(DIR.replace('\\test_data', "\\db_config.cfg"))

        http = config.get('PROXIES', 'http')
        https = config.get('PROXIES', 'https')


        r = requests.get(download_url, proxies={'http':http, 'https':https})

        with open(zip_path, 'wb') as f:
            f.write(r.content)
    gdb = os.path.join(os.path.dirname(zip_path), 'lion/lion.gdb')
    if not os.path.isfile(gdb):
        print ('extracting\n\n')
        print(os.path.dirname(zip_path))
        print('\n\n')
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(os.path.dirname(zip_path))


def clean_up_feature_class():
    zip_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'test_data/nyclion_21d.zip')
    fldr = os.path.dirname(zip_path)+'/lion'
    gdb = "lion.gdb"
    print ('Deleting any existing gdb')
    # os.remove(os.path.join(fldr, gdb))
    os.rmdir(fldr)

# def set_up_fc_and_shapefile():
#     """
#     Builds file gdb with a feature class with sample data
#     Builds a shapefile for import
#     :return:
#     """
#     fldr = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'test_data')
#     set_up_feature_class()
#
#     print 'Deleting any existing shp'
#     shp = 'test_feature_class.shp'
#     Delete_management(os.path.join(fldr, shp))
#
#     print 'Building shapefile'
#     FeatureClassToShapefile_conversion(["test_feature_class"], fldr)
#
#
# def clean_up_fc_and_shapefile():
#     fldr = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'test_data')
#     gdb = "fGDB.gdb"
#     shp = 'test_feature_class.shp'
#
#     Delete_management(os.path.join(fldr, gdb))
#     Delete_management(os.path.join(fldr, shp))
#     print 'Deleted ESRI sample files'
#
#
def set_up_shapefile():
    data = {
        'gid': {0: 1, 1: 2},
        'WKT': {0: 'POINT(-73.88782477721676 40.75343453961836)', 1: 'POINT(-73.88747073046778 40.75149365677327)'},
        'some_value': {0: 'test1', 1: 'test2'}
    }
    df = pd.DataFrame(data)
    df.to_csv(os.path.join(DIR, "sample.csv"), index=False)
    fle = os.path.join(DIR, "sample.csv")

    cmd = f'''ogr2ogr -f "ESRI Shapefile" {DIR}\\test.shp -dialect sqlite -sql 
    "SELECT gid, GeomFromText(WKT, 4326), some_value FROM sample" {fle}'''
    os.system(cmd.replace('\n', ' '))
    print ('Sample shapefile ready...')

    # Add shpfile to zip for testing
    with ZipFile(os.path.join(DIR, 'test.zip'), 'w') as z:
        for ext in ('shp', 'dbf', 'shx', 'prj'):
            _fle = os.path.join(DIR,f'test.{ext}')

            if os.path.isfile(os.path.join(DIR,f'test.{ext}')):
                filePath = os.path.join(DIR, f'test.{ext}')
                z.write(filePath, os.path.basename(filePath))
    print('Sample zipped shapefile ready...')


def clean_up_shapefile():
    fldr = os.path.join(os.path.dirname(os.path.abspath(__file__)))
    for ext in ('shp', 'dbf', 'shx', 'prj', 'zip'):
        _fle = f'{fldr}\\test_data\\test.{ext}'
        if os.path.isfile(_fle):
            os.remove(_fle)

    print ('Deleting any existing shp')
    # Delete_management(os.path.join(fldr, shp))


def set_up_schema(db, ms_schema='dbo', pg_schema='working'):
    if db.type == 'MS':
        db.query(f"""
            IF NOT EXISTS (
                SELECT  schema_name
                FROM    information_schema.schemata
                WHERE   schema_name = '{ms_schema}' 
            ) 
             
            BEGIN
            EXEC sp_executesql N'CREATE SCHEMA {ms_schema}'
            END
        """)
    if db.type == 'PG':
        db.query(f"""
            create schema if not exists {pg_schema};
        """)


def clean_up_schema(db, schema):
    if db.type == 'PG':
        c = ' cascade'
    else:
        c = ''
    db.query("DROP SCHEMA IF EXISTS {}{};".format(schema, c))

def clean_up_shp(file_path):
    for ext in ('.shp', '.dbf', '.shx', '.prj'):
        clean_up_file(file_path.replace('.shp', ext))


def clean_up_file(file_path):
    if os.path.isfile(file_path):
        os.remove(file_path)
        print ('%s file removed\n' % os.path.basename(file_path))


def set_up_xls():
    xls_file1 = os.path.join(DIR, 'test_xls.xls')
    if os.path.isfile(xls_file1):
        clean_up_file(xls_file1)

    test_df1 = pd.DataFrame({'a': {0: 1, 1: 2, 2:3}, 'b': {0: 3, 1: 4, 2:5}, 'Unnamed: 0': {0: 0, 1: 1, 2:6}})
    test_df1.to_excel(os.path.join(DIR, 'test_xls.xls'), index=False)
    print ('%s created\n' % os.path.basename(xls_file1))

    xls_file2 = os.path.join(DIR, 'test_xls_with_sheet.xls')
    if os.path.isfile(xls_file2):
        clean_up_file(xls_file2)

    test_df2 = pd.DataFrame({'a': {0: 1, 1: 2}, 'b': {0: 3, 1: 4}, 'Unnamed: 0': {0: 0, 1: 1}})

    test_df2.to_excel(os.path.join(DIR, 'test_xls_with_sheet.xls'), sheet_name='AnotherSheet', index=False)
    w = copy(open_workbook(xls_file2))
    Sheet2 = w.add_sheet('Sheet2')
    col, row = 0, 0
    for c in test_df2.columns:
        Sheet2.write(row, col, c)
        row += 1
        for r in test_df2[c]:
            Sheet2.write(row, col, r)
            row += 1
        col += 1
        row = 0
    w.save(xls_file2)
    print ('%s created\n' % os.path.basename(xls_file2))


    # set up xls files for kwargs testing
    wb = openpyxl.Workbook()
    sheet = wb.active
    c1 = sheet.cell(row=1, column=1)
    c1.value = "skip me"
    for c in range(1,4):
        cell = sheet.cell(row=2, column=c)
        cell.value = f'header {c}'
        cell2 = sheet.cell(row=3, column=c)
        cell2.value = c
        cell3 = sheet.cell(row=3, column=c)
        cell3.value = c+1

    wb.save(DIR + "\\xls_kwargs_test.xls")
    _df = pd.read_excel(DIR + "\\xls_kwargs_test.xls")
    _df.to_excel(DIR + "\\xls_kwargs_test.xlsx", index=False)

